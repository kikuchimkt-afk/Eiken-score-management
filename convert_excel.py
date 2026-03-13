import openpyxl
import json

wb = openpyxl.load_workbook('英検2025年第2回2次までの結果.xlsx')
ws = wb.active

excel_headers = [cell.value for cell in ws[1]]

# Fields we want in our JSON
fields = [
    'year', 'session', 'grade', 'schoolYear', 'name', 'exemption',
    'primaryResult', 'secondaryResult', 'overallResult',
    'bandPrimary', 'bandSecondary',
    'readingCSE', 'listeningCSE', 'writingCSE', 'primaryCSETotal',
    'speakingCSE', 'totalCSE',
    'readingCEFR', 'listeningCEFR', 'writingCEFR', 'speakingCEFR', 'overallCEFR',
    'readingQ1', 'readingQ2', 'readingQ3', 'readingQ4',
    'readingTotal', 'readingRate',
    'listeningQ1', 'listeningQ2', 'listeningQ3',
    'listeningTotal', 'listeningRate',
    'writingS1Content', 'writingS1Structure', 'writingS1Vocab', 'writingS1Grammar',
    'writingS2Content', 'writingS2Structure', 'writingS2Vocab', 'writingS2Grammar',
    'writingScore', 'writingRate',
    'speakingReading', 'speakingQA', 'speakingAttitude', 'speakingScore'
]


def safe_val(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if v in ('', '－', '-', '*'):
            return None
        if v.startswith('='):
            return None
        try:
            return int(v)
        except ValueError:
            try:
                return float(v)
            except ValueError:
                return v
    return v


data = []
for r in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx in range(min(len(excel_headers), len(fields))):
        field = fields[col_idx]
        cell_val = ws.cell(row=r, column=col_idx + 1).value
        row_data[field] = safe_val(cell_val)

    # Compute formula values
    # readingTotal
    rq = [row_data.get('readingQ' + str(i)) for i in range(1, 5)]
    rq_nums = [x for x in rq if isinstance(x, (int, float))]
    if rq_nums and row_data.get('exemption') != '有':
        row_data['readingTotal'] = sum(rq_nums)

    # readingRate
    g = row_data.get('grade', '')
    if row_data.get('readingTotal') is not None and isinstance(row_data['readingTotal'], (int, float)):
        denoms = {'5級': 25, '4級': 35, '3級': 30, '準2級': 29, '準2級+': 29, '2級': 31}
        d = denoms.get(g, 30)
        row_data['readingRate'] = round(row_data['readingTotal'] / d, 4)

    # listeningTotal
    lq = [row_data.get('listeningQ' + str(i)) for i in range(1, 4)]
    lq_nums = [x for x in lq if isinstance(x, (int, float))]
    if lq_nums and row_data.get('exemption') != '有':
        row_data['listeningTotal'] = sum(lq_nums)

    # listeningRate
    if row_data.get('listeningTotal') is not None and isinstance(row_data['listeningTotal'], (int, float)):
        d = 25 if g == '5級' else 30
        row_data['listeningRate'] = round(row_data['listeningTotal'] / d, 4)

    # writingScore
    ws_fields = [
        'writingS1Content', 'writingS1Structure', 'writingS1Vocab', 'writingS1Grammar',
        'writingS2Content', 'writingS2Structure', 'writingS2Vocab', 'writingS2Grammar'
    ]
    ws_nums = [row_data.get(f) for f in ws_fields if isinstance(row_data.get(f), (int, float))]
    if ws_nums and g not in ('4級', '5級'):
        row_data['writingScore'] = sum(ws_nums)

    # speakingScore
    sp_fields = ['speakingReading', 'speakingQA', 'speakingAttitude']
    sp_nums = [row_data.get(f) for f in sp_fields if isinstance(row_data.get(f), (int, float))]
    if sp_nums:
        row_data['speakingScore'] = sum(sp_nums)

    data.append(row_data)

# Write as data.js
with open('data.js', 'w', encoding='utf-8') as f:
    f.write('const INITIAL_DATA = ')
    json.dump(data, f, ensure_ascii=False, indent=2)
    f.write(';\n')

print("data.js 生成完了: {}件のレコード".format(len(data)))
for i in [0, 1, 80, 81]:
    d = data[i]
    print("  [{}] {} ({}, {}): R={}, L={}, W={}, S={}".format(
        i, d['name'], d['grade'], d['session'],
        d.get('readingTotal'), d.get('listeningTotal'),
        d.get('writingScore'), d.get('speakingScore')
    ))
